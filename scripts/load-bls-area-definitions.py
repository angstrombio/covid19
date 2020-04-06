import argparse
import coronadb
import psycopg2
from psycopg2 import extras
from openpyxl import load_workbook


def load(source):
    with psycopg2.connect(dbname=coronadb.database, port=coronadb.port, user=coronadb.user, host=coronadb.host, password=coronadb.password) as db:
        clear_data(db)
        load_data(db, source)


def clear_data(db):
    with db.cursor() as cursor:
        # noinspection SqlWithoutWhere
        cursor.execute("DELETE FROM covid19.bls_areas")

    db.commit()
    print("Cleaned database")


def load_data(db, source):
    wb = load_workbook(filename=source, read_only=True)
    sheet = wb['area_definitions_m2019']
    count = 1
    num_rows = sheet.max_row-1
    all_areas = []
    all_counties = []
    existing_area_ids = []

    with db.cursor() as cursor:
        for row in sheet.iter_rows(min_row=2, max_row=0):
            print(str(count) + " / " + str(num_rows), end='\r')
            fips = row[0].value
            state = row[1].value
            state_abbr = row[2].value
            msa_code = row[3].value
            if isinstance(msa_code, int):
                msa_code = str(msa_code)
            else:
                msa_code = msa_code.strip()
            msa_name = row[4].value
            county_code = row[5].value
            township_code = row[6].value
            county_name = row[7].value
            if len(msa_code) == 5:
                area_type = '4'
            else:
                area_type = '6'
            if msa_code not in existing_area_ids:
                all_areas.append((msa_code, area_type, msa_name))
                existing_area_ids.append(msa_code)
            all_counties.append((msa_code, fips, state, state_abbr, county_code, township_code, county_name, fips+county_code))

            count += 1

        print()
        print("Executing SQL")
        extras.execute_values(
            cursor,
            "INSERT INTO covid19.bls_areas (area_id, area_type, msa_name) VALUES %s",
            all_areas)
        extras.execute_values(
            cursor,
            "INSERT INTO covid19.bls_area_counties (area_id, fips, state, state_abbr, county_code, township_code, county_or_township_name, fips_stcou) VALUES %s",
            all_counties)

    print()
    db.commit()


parser = argparse.ArgumentParser(description='Script to load BLS area definitions into the database')
parser.add_argument("--source", required=True, type=str, help="File to load")
args = parser.parse_args()
load(args.source)
