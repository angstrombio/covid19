import argparse
import coronadb
import psycopg2
from psycopg2 import extras
from openpyxl import load_workbook


def load(source):
    with psycopg2.connect(dbname=coronadb.database, port=coronadb.port, user=coronadb.user, host=coronadb.host, password=coronadb.password) as db:
        load_data(db, source)


def load_data(db, source):
    wb = load_workbook(filename=source, read_only=True)
    if 'All May 2019 Data' in wb:
        sheet = wb['All May 2019 Data']
    else:
        sheet = wb['BOS_M2019_dl']
    count = 1
    num_rows = sheet.max_row-1
    all_values = []
    batch_size = 10000

    with db.cursor() as cursor:
        for row in sheet.iter_rows(min_row=2, max_row=0):
            print(str(count) + " / " + str(num_rows), end='\r')
            area = row[0].value
            area_title = row[1].value
            area_type = row[2].value
            occ_code = row[7].value
            occ_title = row[8].value
            o_group = row[9].value
            tot_emp = row[10].value
            if tot_emp == '**':
                tot_emp = None

            all_values.append((area, area_title, area_type, occ_code, occ_title, o_group, tot_emp))

            if len(all_values) >= batch_size:
                insert(cursor, all_values)
                all_values = []

            count += 1

        if len(all_values) > 0:
            insert(cursor, all_values)

    print()
    db.commit()


def insert(cursor, all_values):
    print()
    print("Inserting " + str(len(all_values)))
    extras.execute_values(
        cursor,
        "INSERT INTO covid19.bls_oes (area, area_title, area_type, occ_code, occ_title, o_group, tot_emp) VALUES %s",
        all_values)


parser = argparse.ArgumentParser(description='Script to load BLS metro area OES datainto the database')
parser.add_argument("--source", required=True, type=str, help="File to load")
args = parser.parse_args()
load(args.source)
