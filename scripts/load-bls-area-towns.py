import argparse
import coronadb
import psycopg2
from psycopg2 import extras
from openpyxl import load_workbook


def load(source):
    with psycopg2.connect(dbname=coronadb.database, port=coronadb.port, user=coronadb.user, host=coronadb.host, password=coronadb.password) as db:
        load_data(db, source)

def load_data(db, source):
    wb = load_workbook(filename=source, read_only=True)
    sheet = wb['load bls_area_towns']
    count = 1
    num_rows = sheet.max_row-1
    all_data = []

    with db.cursor() as cursor:
        cursor.execute("select state_code, state_name from covid19.census_state_codes")
        rows = cursor.fetchall()
        states = {}
        for row in rows:
            states[row[1]] = row[0]

        for row in sheet.iter_rows(min_row=2, max_row=0):
            print(str(count) + " / " + str(num_rows), end='\r')
            area_id = row[0].value
            state_code = states[row[1].value]
            county_code = row[2].value
            township_code = row[3].value
            town_or_county_name = row[4].value
            pct = row[5].value
            fips_stcou = state_code + county_code

            all_data.append((area_id, state_code, county_code, township_code, town_or_county_name, pct, fips_stcou))
            count += 1

        print()
        print("Executing SQL")
        extras.execute_values(
            cursor,
            "INSERT INTO covid19.bls_area_towns (area_id, state_code, county_code, township_code, town_or_county_name, pct_of_msa_by_population, fips_stcou) VALUES %s",
            all_data)

    print()
    db.commit()


parser = argparse.ArgumentParser(description='Script to load BLS data into the database')
parser.add_argument("--source", required=True, type=str, help="File to load")
args = parser.parse_args()
load(args.source)
